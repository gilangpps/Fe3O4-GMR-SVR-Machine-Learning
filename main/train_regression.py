"""
GMR Fe3O4 Sensor - Regression Model Training
Models: SVR, KNN Regressor, Random Forest Regressor
Input: Delta-B (mT) from processed sheet, all iterations
Target: Continuous concentration (mg/mL)
"""

import os
import json
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import seaborn as sns
import joblib

from sklearn.pipeline import Pipeline
from sklearn.svm import SVR
from sklearn.neighbors import KNeighborsRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split, KFold, cross_val_score
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import (
    mean_absolute_error,
    mean_squared_error,
    r2_score,
    mean_absolute_percentage_error,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
sns.set_theme(style="whitegrid", context="talk")

# ─── CONFIG ──────────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))

DATA_PATH = os.path.join(_HERE, "090526_fe3o4-data-gmr.xlsx")
SHEET_NAME = "Data_Acquisition_Processed"

BASE_OUT_DIR = os.path.join(_HERE, "output_results", "regression")
MODEL_DIR = os.path.join(_HERE, "models", "regression")

PLOT_DIR = os.path.join(BASE_OUT_DIR, "plots")
INDIVIDUAL_PLOT_DIR = os.path.join(PLOT_DIR, "individual")
COMPARISON_PLOT_DIR = os.path.join(PLOT_DIR, "comparison")
EXCEL_DIR = os.path.join(BASE_OUT_DIR, "excel")
JSON_DIR = os.path.join(BASE_OUT_DIR, "json")

for d in [
    MODEL_DIR,
    PLOT_DIR,
    INDIVIDUAL_PLOT_DIR,
    COMPARISON_PLOT_DIR,
    EXCEL_DIR,
    JSON_DIR,
]:
    os.makedirs(d, exist_ok=True)

RANDOM_STATE = 42
TEST_SIZE = 0.2
CV_FOLDS = 5

CONCENTRATIONS = [5, 10, 20, 30, 40, 50]
COL_GROUPS = {
    5:  [1, 2, 3, 4, 5],
    10: [6, 7, 8, 9, 10],
    20: [11, 12, 13, 14, 15],
    30: [16, 17, 18, 19, 20],
    40: [21, 22, 23, 24, 25],
    50: [26, 27, 28, 29, 30],
}

COLORS = {
    "SVR": "#2563EB",
    "KNN": "#16A34A",
    "RandomForest": "#DC2626",
}

# ─── UTILITIES ───────────────────────────────────────────────────────────────
def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

def style_header(cell, bg="7C2D12"):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

def style_cell(cell):
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

def load_data(path, sheet):
    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    data = raw.iloc[2:].reset_index(drop=True)

    records = []
    for conc, cols in COL_GROUPS.items():
        for c in cols:
            col_data = pd.to_numeric(data.iloc[:, c], errors="coerce").dropna().values
            for val in col_data:
                records.append({"delta_B_mT": float(val), "concentration": float(conc)})

    return pd.DataFrame(records)

def build_models():
    return {
        "SVR": Pipeline(
            steps=[
                ("scaler", StandardScaler()),
                ("model", SVR(kernel="rbf", C=100, gamma="scale", epsilon=0.5)),
            ]
        ),
        "KNN": Pipeline(
            steps=[
                ("scaler", StandardScaler()),
                ("model", KNeighborsRegressor(n_neighbors=7, metric="euclidean")),
            ]
        ),
        "RandomForest": Pipeline(
            steps=[
                ("model", RandomForestRegressor(n_estimators=200, max_depth=None, random_state=RANDOM_STATE)),
            ]
        ),
    }

def save_metrics_json(payload, output_path):
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=4, ensure_ascii=False)

# ─── EXCEL OUTPUT ─────────────────────────────────────────────────────────────
def write_metrics_excel(all_metrics, cv_results, residual_dict, output_path):
    wb = Workbook()

    # Sheet 1: Summary Metrics
    ws = wb.active
    ws.title = "Summary Metrics"
    headers = ["Model", "MAE (mg/mL)", "MSE", "RMSE (mg/mL)", "R²", "MAPE (%)", "CV Mean R²", "CV Std R²"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(1, col, h))

    for row, (name, m) in enumerate(all_metrics.items(), 2):
        vals = [name, m["mae"], m["mse"], m["rmse"], m["r2"], m["mape"], m["cv_mean"], m["cv_std"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row, col, round(v, 6) if isinstance(v, float) else v)
            style_cell(c)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Sheet 2: Residuals per model
    for model_name, res_df in residual_dict.items():
        ws2 = wb.create_sheet(f"Residuals_{model_name}")
        for col, h in enumerate(["Actual (mg/mL)", "Predicted (mg/mL)", "Residual", "Abs Residual"], 1):
            style_header(ws2.cell(1, col, h))
        for row, (_, r) in enumerate(res_df.iterrows(), 2):
            for col, v in enumerate([r["actual"], r["predicted"], r["residual"], r["abs_residual"]], 1):
                c = ws2.cell(row, col, round(float(v), 6))
                style_cell(c)
        for col in range(1, 5):
            ws2.column_dimensions[get_column_letter(col)].width = 20

    # Sheet 3: CV Results
    ws3 = wb.create_sheet("Cross_Validation")
    style_header(ws3.cell(1, 1, "Model"))
    for fold in range(1, CV_FOLDS + 1):
        style_header(ws3.cell(1, fold + 1, f"Fold {fold} R²"))
    style_header(ws3.cell(1, CV_FOLDS + 2, "Mean R²"))
    style_header(ws3.cell(1, CV_FOLDS + 3, "Std R²"))

    for row, (name, scores) in enumerate(cv_results.items(), 2):
        ws3.cell(row, 1, name).font = Font(bold=True, name="Arial")
        for fold, s in enumerate(scores, 2):
            c = ws3.cell(row, fold, round(float(s), 6))
            style_cell(c)
        ws3.cell(row, CV_FOLDS + 2, round(float(np.mean(scores)), 6))
        ws3.cell(row, CV_FOLDS + 3, round(float(np.std(scores)), 6))

    for col in range(1, CV_FOLDS + 4):
        ws3.column_dimensions[get_column_letter(col)].width = 15

    wb.save(output_path)
    print(f"  [Excel] Saved -> {output_path}")

# ─── VISUALIZATION ────────────────────────────────────────────────────────────
def plot_regression_individual(model_name, metrics, y_test, y_pred, residuals, cv_scores, output_dir):
    ensure_dir(output_dir)

    color = COLORS[model_name]
    fig = plt.figure(figsize=(18, 12))
    fig.suptitle(
        f"{model_name} Regression — GMR Fe₃O₄ Sensor (ΔB, mT)",
        fontsize=15,
        fontweight="bold",
        y=0.98,
    )
    gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.45, wspace=0.4)

    # 1. Actual vs Predicted
    ax1 = fig.add_subplot(gs[0, 0])
    ax1.scatter(y_test, y_pred, alpha=0.45, s=22, color=color, edgecolors="none")
    lims = [min(y_test.min(), y_pred.min()) - 1, max(y_test.max(), y_pred.max()) + 1]
    ax1.plot(lims, lims, "k--", lw=1.5, label="Ideal (y=x)")
    ax1.set_xlabel("Actual Concentration (mg/mL)")
    ax1.set_ylabel("Predicted Concentration (mg/mL)")
    ax1.set_title(f"Actual vs Predicted\nR² = {metrics['r2']:.4f}", fontweight="bold")
    ax1.legend(fontsize=8)

    # 2. Residuals vs Actual
    ax2 = fig.add_subplot(gs[0, 1])
    ax2.scatter(y_test, residuals, alpha=0.45, s=22, color=color, edgecolors="none")
    ax2.axhline(0, color="black", lw=1.5, ls="--")
    ax2.set_xlabel("Actual Concentration (mg/mL)")
    ax2.set_ylabel("Residual (mg/mL)")
    ax2.set_title("Residuals vs Actual", fontweight="bold")

    # 3. Residual distribution
    ax3 = fig.add_subplot(gs[0, 2])
    ax3.hist(residuals, bins=30, color=color, alpha=0.75, edgecolor="white")
    ax3.axvline(0, color="black", lw=1.5, ls="--")
    ax3.set_xlabel("Residual (mg/mL)")
    ax3.set_ylabel("Count")
    ax3.set_title("Residual Distribution", fontweight="bold")

    # 4. Summary metrics
    ax4 = fig.add_subplot(gs[1, 0])
    metric_names = ["MAE", "RMSE", "R²", "MAPE (%)"]
    metric_vals = [metrics["mae"], metrics["rmse"], metrics["r2"], metrics["mape"]]
    bars = ax4.bar(metric_names, metric_vals, color=color, alpha=0.85, edgecolor="white")
    ax4.set_title("Overall Metrics Summary", fontweight="bold")
    ax4.set_ylabel("Value")
    for bar, v in zip(bars, metric_vals):
        ax4.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + max(metric_vals) * 0.01,
            f"{v:.4f}",
            ha="center",
            va="bottom",
            fontsize=8,
        )

    # 5. Absolute error per concentration class
    ax5 = fig.add_subplot(gs[1, 1])
    groups = {c: [] for c in CONCENTRATIONS}
    for actual, pred in zip(y_test, y_pred):
        c_key = int(round(actual))
        if c_key in groups:
            groups[c_key].append(abs(actual - pred))

    box_data = [groups[c] for c in CONCENTRATIONS]
    bp = ax5.boxplot(box_data, patch_artist=True, labels=[str(c) for c in CONCENTRATIONS])
    for patch in bp["boxes"]:
        patch.set_facecolor(color)
        patch.set_alpha(0.7)
    ax5.set_xlabel("Concentration (mg/mL)")
    ax5.set_ylabel("Absolute Error (mg/mL)")
    ax5.set_title("Error per Concentration Class", fontweight="bold")

    # 6. CV R² across folds
    ax6 = fig.add_subplot(gs[1, 2])
    ax6.bar([f"Fold {i+1}" for i in range(len(cv_scores))], cv_scores, color=color, alpha=0.85, edgecolor="white")
    ax6.axhline(np.mean(cv_scores), color="black", ls="--", lw=1.5, label=f"Mean = {np.mean(cv_scores):.4f}")
    ax6.set_title(f"{CV_FOLDS}-Fold CV R²", fontweight="bold")
    ax6.set_ylabel("R²")
    ax6.legend(fontsize=8)
    ax6.tick_params(axis="x", rotation=30)

    fig.tight_layout(rect=[0, 0, 1, 0.96])
    out_path = os.path.join(output_dir, f"regression_{model_name.lower()}_analysis.png")
    plt.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  [Plot] Saved -> {out_path}")

def plot_regression_comparison(all_metrics, cv_results, y_test, preds, output_dir):
    ensure_dir(output_dir)

    fig = plt.figure(figsize=(22, 14))
    fig.suptitle(
        "Regression Comparison — SVR vs KNN vs Random Forest\nGMR Fe₃O₄ Sensor (ΔB, mT)",
        fontsize=15,
        fontweight="bold",
        y=0.99,
    )
    gs = gridspec.GridSpec(3, 3, figure=fig, hspace=0.55, wspace=0.4)

    model_names = list(all_metrics.keys())
    color_map = [COLORS[n] for n in model_names]

    # Row 0: Actual vs Predicted scatter per model
    for idx, (name, color) in enumerate(zip(model_names, color_map)):
        ax = fig.add_subplot(gs[0, idx])
        y_pred = preds[name]
        ax.scatter(y_test, y_pred, alpha=0.35, s=15, color=color, edgecolors="none")
        lims = [4, 52]
        ax.plot(lims, lims, "k--", lw=1.2)
        ax.set_xlabel("Actual (mg/mL)", fontsize=8)
        ax.set_ylabel("Predicted (mg/mL)", fontsize=8)
        ax.set_title(f"{name} — Actual vs Predicted\nR² = {all_metrics[name]['r2']:.4f}", fontweight="bold", fontsize=10)

    # Row 1 left: Metric bar comparison
    ax_bar = fig.add_subplot(gs[1, :2])
    metric_keys = ["mae", "rmse", "r2", "mape"]
    metric_labels = ["MAE (mg/mL)", "RMSE (mg/mL)", "R²", "MAPE (%)"]
    x = np.arange(len(metric_labels))
    width = 0.25
    for i, (name, color) in enumerate(zip(model_names, color_map)):
        vals = [all_metrics[name][k] for k in metric_keys]
        ax_bar.bar(x + (i - 1) * width, vals, width, label=name, color=color, alpha=0.85)
    ax_bar.set_xticks(x)
    ax_bar.set_xticklabels(metric_labels)
    ax_bar.set_title("Metric Comparison (All Models)", fontweight="bold")
    ax_bar.set_ylabel("Value")
    ax_bar.legend()

    # Row 1 right: CV R² boxplot
    ax_cv = fig.add_subplot(gs[1, 2])
    cv_data = [cv_results[n] for n in model_names]
    bp = ax_cv.boxplot(cv_data, patch_artist=True, labels=model_names)
    for patch, color in zip(bp["boxes"], color_map):
        patch.set_facecolor(color)
        patch.set_alpha(0.7)
    ax_cv.set_title(f"{CV_FOLDS}-Fold CV R²", fontweight="bold")
    ax_cv.set_ylabel("R²")

    # Row 2: Residual distributions
    for idx, (name, color) in enumerate(zip(model_names, color_map)):
        ax = fig.add_subplot(gs[2, idx])
        res = y_test - preds[name]
        ax.hist(res, bins=30, color=color, alpha=0.75, edgecolor="white")
        ax.axvline(0, color="black", lw=1.5, ls="--")
        ax.set_xlabel("Residual (mg/mL)", fontsize=8)
        ax.set_ylabel("Count", fontsize=8)
        ax.set_title(f"{name} — Residual Distribution", fontweight="bold", fontsize=10)

    fig.tight_layout(rect=[0, 0, 1, 0.96])
    out_path = os.path.join(output_dir, "regression_comparison_all_models.png")
    plt.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  [Plot] Saved -> {out_path}")

# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  GMR Fe3O4 — Regression Training")
    print("=" * 60)

    df = load_data(DATA_PATH, SHEET_NAME)
    print(f"  Dataset: {len(df)} samples")

    X = df[["delta_B_mT"]].values
    y = df["concentration"].values

    X_train, X_test, y_train, y_test = train_test_split(
        X,
        y,
        test_size=TEST_SIZE,
        random_state=RANDOM_STATE,
    )
    print(f"  Train: {len(X_train)} | Test: {len(X_test)}")

    models = build_models()
    kf = KFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)

    all_metrics = {}
    cv_results = {}
    residual_dict = {}
    preds = {}

    for name, model in models.items():
        print(f"\n  Training {name}...")
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)

        mae = mean_absolute_error(y_test, y_pred)
        mse = mean_squared_error(y_test, y_pred)
        rmse = np.sqrt(mse)
        r2 = r2_score(y_test, y_pred)
        mape = mean_absolute_percentage_error(y_test, y_pred) * 100
        cv_sc = cross_val_score(model, X, y, cv=kf, scoring="r2")

        all_metrics[name] = {
            "mae": float(mae),
            "mse": float(mse),
            "rmse": float(rmse),
            "r2": float(r2),
            "mape": float(mape),
            "cv_mean": float(np.mean(cv_sc)),
            "cv_std": float(np.std(cv_sc)),
        }
        cv_results[name] = cv_sc
        preds[name] = y_pred

        res = y_test - y_pred
        residual_dict[name] = pd.DataFrame(
            {
                "actual": y_test,
                "predicted": y_pred,
                "residual": res,
                "abs_residual": np.abs(res),
            }
        )

        print(f"    MAE  : {mae:.4f} mg/mL")
        print(f"    RMSE : {rmse:.4f} mg/mL")
        print(f"    R²   : {r2:.4f}")
        print(f"    MAPE : {mape:.2f} %")
        print(f"    CV R²: {np.mean(cv_sc):.4f} ± {np.std(cv_sc):.4f}")

        model_dir = os.path.join(MODEL_DIR, name)
        ensure_dir(model_dir)
        model_path = os.path.join(model_dir, f"regressor_{name.lower()}.pkl")
        joblib.dump(model, model_path)
        print(f"    [Model] Saved -> {model_path}")

        plot_regression_individual(
            model_name=name,
            metrics=all_metrics[name],
            y_test=y_test,
            y_pred=y_pred,
            residuals=res,
            cv_scores=cv_sc,
            output_dir=INDIVIDUAL_PLOT_DIR,
        )

    excel_path = os.path.join(EXCEL_DIR, "regression_metrics.xlsx")
    write_metrics_excel(all_metrics, cv_results, residual_dict, excel_path)

    json_path = os.path.join(JSON_DIR, "regression_metrics.json")
    save_metrics_json(all_metrics, json_path)
    print(f"  [JSON] Saved -> {json_path}")

    print("\n  Generating comparison plot...")
    plot_regression_comparison(
        all_metrics=all_metrics,
        cv_results=cv_results,
        y_test=y_test,
        preds=preds,
        output_dir=COMPARISON_PLOT_DIR,
    )

    print("\n" + "=" * 60)
    print("  Regression training complete.")
    print("=" * 60)

if __name__ == "__main__":
    main()